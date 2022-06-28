from tkinter import *
from tkinter import messagebox
from PIL import ImageTk, Image
from openpyxl import load_workbook
import words

atm = 1
guessnum = 1
btnState = False
word = "start"
running = False
minutes, seconds = 0, 0
str = "time"
temp = "time"
noL = 5
n = "name"


def gamePlay5(t):
    t.destroy()
    name = Tk()
    name.config(bg="#292F3F")
    name.geometry("275x90")
    name.title("Enter name")
    Label(name,text="Enter your name : ").place(x=20,y=20)
    player_name = StringVar()
    e = Entry(textvariable=player_name)
    e.place(y=20,x=130)
    global n
    n = player_name.get()
    btn = Button(text = "Continue",command=lambda : name.destroy())
    btn.place(x=110,y=50)
    name.mainloop()

    root = Tk()
    root.title("WORDLE")
    root.geometry("330x530")
    root.iconbitmap('5.ico')

    global word
    word = words.get_word()
    print(word)

    def saveScore(n,a,m,s):
        wb = load_workbook('score.xlsx')
        ws = wb.worksheets[0]
        r = ws.max_row
        ws.cell(row=r+1,column=1,value=n)
        ws.cell(row=r+1,column=2,value=a)
        ws.cell(row=r+1,column=3,value=m)
        ws.cell(row=r+1,column=4,value=s)
        wb.save('score.xlsx')

    def start():
        global running
        if not running:
            update()
            running = True

    def reset():
        global running
        if running:
            # cancel updating of time using after_cancel()
            stopwatch_label.after_cancel(update_time)
            running = False
        # set variables back to zero
        global minutes, seconds
        minutes, seconds = 0, 0
        # set label back to zero
        stopwatch_label.config(text='00:00')

    def pause():
        global running
        if running:
            # cancel updating of time using after_cancel()
            stopwatch_label.after_cancel(update_time)
            running = False

    def update():
        # update seconds with (addition) compound assignment operator
        global minutes, seconds
        seconds += 1
        if seconds == 60:
            minutes += 1
            seconds = 0
        # format time to include leading zeros
        minutes_string = f'{minutes}' if minutes > 9 else f'0{minutes}'
        seconds_string = f'{seconds}' if seconds > 9 else f'0{seconds}'
        # update timer label after 1000 ms (1 second)
        stopwatch_label.config(text=minutes_string + ':' + seconds_string)
        # after each second (1000 milliseconds), call update function
        # use update_time variable to cancel or pause the time using after_cancel
        global update_time
        update_time = stopwatch_label.after(1000, update)

    GREEN = "#007d21"
    YELLOW = "#e2e600"
    BLACK = "#000000"
    WHITE = "#FFFFFF"

    root.config(bg="#CECCBE")
    stopwatch_label = Label(text='00:00', bg="#CECCBE", fg="black", font=('Arial', 10))
    stopwatch_label.place(x=20, y=20)
    reset()
    start()

    # setting switch function:
    def switch():
        global btnState
        if btnState:
            btn.config(image=offImg, bg="#CECCBE", activebackground="#CECCBE")
            root.config(bg="#CECCBE")
            f.config(bg="#CECCBE")
            stopwatch_label.config(bg="#CECCBE", fg="black")
            txt.config(text="Dark Mode: OFF", bg="#CECCBE")
            btnState = False
        else:
            btn.config(image=onImg, bg="#2B2B2B", activebackground="#2B2B2B")
            root.config(bg="#2B2B2B")
            f.config(bg="#2B2B2B")
            stopwatch_label.config(bg="#2B2B2B", fg="white")

            txt.config(text="Dark Mode: ON", bg="#2B2B2B")
            btnState = True

    # loading the switch images:
    image = Image.open('dark.png')
    image = image.resize((40, 40), Image.ANTIALIAS)
    onImg = ImageTk.PhotoImage(image)

    image2 = Image.open('light.png')
    image2 = image2.resize((40, 40), Image.ANTIALIAS)
    offImg = ImageTk.PhotoImage(image2)

    # onImg = PhotoImage(file=r"light.png")
    # offImg = PhotoImage(file=r"dark.png")

    # Night mode label:
    txt = Label(root, text="Dark Mode: OFF", font="FixedSys 10", bg="#CECCBE", fg="green")
    txt.place(x=130, y=20)

    # switch widget:
    btn = Button(root, text="OFF", borderwidth=0, command=switch, bg="#CECCBE", activebackground="#CECCBE")
    btn.place(x=260, y=10, width=40, height=35)
    btn.config(image=offImg)

    def getGuess():
        global word
        guess = wordInput.get()
        try:
            for i in range(5):
                a = int(guess[i])
                print(a)
            # a = int(guess)
            messagebox.showinfo("String error", "Please enter alphabets only")

        except:
            global guessnum, atm
            global minutes
            global seconds
            global str
            if guessnum <= 5:
                print(guessnum)
                if len(guess) == 5:
                    if word == guess.lower():
                        for i, letter in enumerate(guess):
                            label = Label(f, text=letter.upper(), height=2, width=5)
                            label.grid(row=guessnum, column=i, padx=10, pady=13)

                            if letter == word[i]:  # if they get the letter right
                                label.config(bg=GREEN, fg=BLACK)

                            if letter in word and not letter == word[
                                i]:  # if the letter is in the word, but not in the right spot
                                label.config(bg=YELLOW, fg=BLACK)

                            if letter not in word:
                                label.config(bg=BLACK, fg=WHITE)
                        atm = guessnum
                        guessnum = guessnum + 1
                        pause()
                        # CORRECT
                        messagebox.showinfo("Correct!", f"Correct! the word was {word.title()}")
                        str = f"The word was: {word}\nYou guessed it in {minutes}:{seconds} and {atm} attempts"
                        saveScore(n,atm,minutes,seconds)
                        root.destroy()



                    else:  # INCORRECT
                        for i, letter in enumerate(guess):
                            label = Label(f, text=letter.upper(), height=2, width=5)
                            label.grid(row=guessnum, column=i, padx=10, pady=13)

                            if letter == word[i]:  # if they get the letter right
                                label.config(bg=GREEN, fg=BLACK)

                            if letter in word and not letter == word[
                                i]:  # if the letter is in the word, but not in the right spot
                                label.config(bg=YELLOW, fg=BLACK)

                            if letter not in word:
                                label.config(bg=BLACK, fg=WHITE)
                        guessnum = guessnum + 1
                else:
                    messagebox.showerror("Character error", "Please use 5 characters in your guess")
                    var.set("")
                var.set("")
            else:
                print(guessnum)
                # guessnum = guessnum - 1
                if word == guess.lower():
                    for i, letter in enumerate(guess):
                        label = Label(f, text=letter.upper(), height=2, width=5)
                        label.grid(row=guessnum, column=i, padx=10, pady=13)

                        if letter == word[i]:  # if they get the letter right
                            label.config(bg=GREEN, fg=BLACK)

                        if letter in word and not letter == word[
                            i]:  # if the letter is in the word, but not in the right spot
                            label.config(bg=YELLOW, fg=BLACK)

                        if letter not in word:
                            label.config(bg=BLACK, fg=WHITE)
                    atm = guessnum
                    guessnum = guessnum + 1
                    pause()
                    # CORRECT
                    messagebox.showinfo("Correct!", f"Correct! the word was {word.title()}")
                    str = f"The word was: {word}\nYou guessed it in {minutes}:{seconds} and {atm} attempts"
                    saveScore(n, atm, minutes, seconds)
                    root.destroy()


                for i, letter in enumerate(guess):
                    label = Label(f, text=letter.upper(), height=2, width=5)
                    label.grid(row=guessnum, column=i, padx=10, pady=13)

                    if letter == word[i]:  # if they get the letter right
                        label.config(bg=GREEN, fg=BLACK)

                    if letter in word and not letter == word[
                        i]:  # if the letter is in the word, but not in the right spot
                        label.config(bg=YELLOW, fg=BLACK)

                    if letter not in word:
                        label.config(bg=BLACK, fg=WHITE)
                guessnum = guessnum + 1
                pause()
                messagebox.showerror(" You lose!", f"You Lose! The word was: {word}")
                var.set("")
                str = f"The word was: {word}\nTime Taken is {minutes}:{seconds}"
                root.destroy()

    f = Frame(root, bg="#CECCBE")
    f.place(x=10, y=55)

    var = StringVar()
    wordInput = Entry(root, textvariable=var, width=25, font="vardana 12")
    wordInput.place(x=50, y=430)

    wordGuessButton = Button(root, text="Guess", font="Vardana 12", width=8, height=1, bg="light green",
                             command=getGuess)
    wordGuessButton.place(x=120, y=460)
    n = player_name.get()
    Label(text=f"Player name : {n}", font="Vardana 10",bg="#CECCBE").place(x=100,y=500)

    root.mainloop()

def returnStr5():
    return str

if __name__ == '__main__':
    root = Tk()
    gamePlay5(root)
    root.mainloop()