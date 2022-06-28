from  tkinter import ttk
from  tkinter import *
from openpyxl import load_workbook

def scoreSort():
    sc = Tk()
    sc.title('High Scores')
    sc.geometry('250x300')
    sc.config(bg='#292F3F')
    game_frame = Frame(sc, bg="#292F3F")
    game_frame.pack()

    style = ttk.Style()
    # style.theme_use("clam")
    style.theme_use("alt")
    # style.theme_use("vista")
    # style.theme_use("default")
    style.configure("Treeview",
                    background="#292F3F",
                    foreground="White",
                    fieldbackground="#292F3F"
                    )
    my_game = ttk.Treeview(game_frame)
    my_game['columns'] = ('Rank', 'player_name', 'Attemps', 'min', 'sec')

    my_game.column("#0", width=0, stretch=NO)
    my_game.column("Rank", anchor=CENTER, width=40)
    my_game.column("player_name", anchor=CENTER, width=80)
    my_game.column("Attemps", anchor=CENTER, width=50)
    my_game.column("min", anchor=CENTER, width=40)
    my_game.column("sec", anchor=CENTER, width=40)

    my_game.heading("#0", text="", anchor=CENTER)
    my_game.heading("Rank", text="Id", anchor=CENTER)
    my_game.heading("player_name", text="Name", anchor=CENTER)
    my_game.heading("Attemps", text="Attemps", anchor=CENTER)
    my_game.heading("min", text="min", anchor=CENTER)
    my_game.heading("sec", text="sec", anchor=CENTER)

    wb = load_workbook('score.xlsx')
    ws = wb.worksheets[3]

    r = ws.max_row
    l1 = []
    for i in range(2, r + 1):
        l2 = []
        for j in range(2, 5):
            c = ws.cell(row=i, column=j).value
            l2.append(c)
        l1.append(l2)
    print(r - 1)
    k = 0
    if (r < 6):
        while (k < r - 1):
            a = 6
            m = 60
            s = 60
            p = 0
            for i in range(len(l1)):
                if int(l1[i][0]) == a:
                    if int(l1[i][1]) == m:
                        if int(l1[i][2]) <= s:
                            a = int(l1[i][0])
                            p = i
                            m = int(l1[i][1])
                            s = int(l1[i][2])
                    elif int(l1[i][1]) < m:
                        a = int(l1[i][0])
                        p = i
                        m = int(l1[i][1])
                        s = int(l1[i][2])
                elif int(l1[i][0]) < a:
                    a = int(l1[i][0])
                    p = i
                    m = int(l1[i][1])
                    s = int(l1[i][2])
                else:
                    continue
            name = ws.cell(row=p + 2, column=1).value
            print(
                f"Rank {k + 1} goes to {name} with completing game in {a} attempts in {m} minutes and {s} seconds")
            my_game.insert(parent='', index='end', text='',
                           values=(f"{k + 1}", f"{name}", f"{a}", f"{m}", f"{s}"))

            l1[p][0] = 7
            k = k + 1
    else:
        while (k < 5):
            a = 6
            m = 60
            s = 60
            p = 0
            for i in range(len(l1)):
                if int(l1[i][0]) == a:
                    if int(l1[i][1]) == m:
                        if int(l1[i][2]) <= s:
                            a = int(l1[i][0])
                            p = i
                            m = int(l1[i][1])
                            s = int(l1[i][2])
                    elif int(l1[i][1]) < m:
                        a = int(l1[i][0])
                        p = i
                        m = int(l1[i][1])
                        s = int(l1[i][2])
                elif int(l1[i][0]) < a:
                    a = int(l1[i][0])
                    p = i
                    m = int(l1[i][1])
                    s = int(l1[i][2])
                else:
                    continue
            name = ws.cell(row=p + 2, column=1).value
            print(
                f"Rank {k + 1} goes to {name} with completing game in {a} attempts in {m} minutes and {s} seconds")
            my_game.insert(parent='', index='end', text='',
                           values=(f"{k + 1}", f"{name}", f"{a}", f"{m}", f"{s}"))

            l1[p][0] = 7
            k = k + 1
    Button(sc, text="Back", bg="light blue", font="Vardana 9", pady=4, height=1, width=6,
           command=lambda: sc.destroy()).place(x=100, y=260)
    my_game.pack()
    sc.mainloop()

scoreSort()